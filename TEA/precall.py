# -*- coding: utf-8 -*-
"""
Created on Thu Aug 20 12:35:46 2020

@author: Melissa Gray

Calculating precision and recall
"""

#%% VARIABLES 7 IMPORTS

import pandas as pd
import numpy as np
import re
import TEA.confusion_matrix as cm
from glob import glob
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
import os.path

from scipy.spatial import distance
from scipy.cluster.hierarchy import linkage, dendrogram, set_link_color_palette
import matplotlib.pyplot as plt


#%% CLASS

class Precall():
    def __init__(self):
        return
    
    def _calculate_precision(self, tp, fp):     # (TP) / (TP+FP)
        precision = 0
        if (tp==0) and (fp==0):
            precision = np.nan
        else:
            precision = (tp) / (tp+fp)
        return precision
    def _calculate_recall(self, tp, fn):        # (TP) / (TP+FN)
        recall = 0
        if (tp==0) and (fn==0):
            recall = np.nan
        else:
            recall = (tp) / (tp+fn)
        return recall
    
    def calculate_precision_and_recall(self, matrices):   # [TP, FN, FP, TN]
        tax_id_precision = {}
        tax_id_recall = {}
        
        for name in matrices:
            for tax_id in matrices[name]:
                if (tax_id not in tax_id_precision):
                    tax_id_precision[tax_id] = {}
                tax_id_precision[tax_id][name] = self._calculate_precision(matrices[name][tax_id][0], matrices[name][tax_id][2])
                if (tax_id not in tax_id_recall):
                    tax_id_recall[tax_id] = {}
                tax_id_recall[tax_id][name] = self._calculate_recall(matrices[name][tax_id][0], matrices[name][tax_id][1])
        for name in matrices:
            for tax_id in tax_id_precision:
                if name not in tax_id_precision[tax_id]:
                    tax_id_precision[tax_id][name] = np.nan
                if name not in tax_id_recall[tax_id]:
                    tax_id_recall[tax_id][name] = np.nan
        return tax_id_precision, tax_id_recall
    
    
    

# For misc stuff
class Misc():
    def __init__(self):
        self.matrix_dict = {}
        self.matrix_tables = {}
        self.saved = {}
        self.cm_truth = ""
        self.input_path = ""
        self.output_path = ""
        self.output_name = ""
        self.true_taxid = []
        return
    
    # GETTERS
    def get_matrx_dict(self):
        return self.matrix_dict
    def get_matrix_tables(self):
        return self.matrix_tables
    def get_saved(self):
        return self.saved
    def get_matrix_names(self):
        names_list = []
        for m in self.matrix_dict:
            names_list.append(m)
        return names_list
    def get_input_path(self):
        return self.input_path
    def get_output_path(self):
        return self.output_path
    def get_output_name(self):
        return self.output_name
    
    # SETTERS
    def set_truth(self, truth):
        self.cm_truth = truth
        return
    def set_input_path(self, ip):
        self.input_path = ip
        return
    def set_output_path(self, op):
        self.output_path = op
        return
    def set_output_name(self, on):
        self.output_name = on
        return
      
    def add_matrix(self, name, matrix):
        self.matrix_dict[name] = matrix
        self.saved[name] = False
        return print("\nAdded as matrix", name)
    
    def remove_matrix(self, name):
        if name in self.matrix_dict:
            del self.matrix_dict[name]
            if name in self.matrix_tables:
                del self.matrix_tables
        else:
            print("There is no matrix by the name \'{}\'".format(name))
        return
    
    def create_table(self, name=""):
        Juice = cm.Confusion(os.path.join(self.input_path, self.cm_truth), "")
        if name == "":
            for name in self.matrix_dict:
                if name not in self.matrix_tables:
                    Juice.set_file_name(os.path.join(self.input_path, name))
                    self.matrix_tables[name] = Juice.create_matrix_table(Juice.reformat_matrix(Juice.add_other_info(self.matrix_dict[name])))
        elif name in self.matrix_dict:
            Juice.set_file_name(os.path.join(self.input_path, name))
            self.matrix_tables[name] = Juice.create_matrix_table(Juice.reformat_matrix(Juice.add_other_info(self.matrix_dict[name])))
        else:
            print("There is no matrix by the name \'{}\'".format(name))
        return
    
    def save_matrices_as_csv(self, file_path):
        Juice = cm.Confusion(os.path.join(self.input_path, self.cm_truth), "")
        for name in self.matrix_dict:
            if self.saved[name] == False:
                Juice.set_file_name(os.path.join(self.input_path, name))
                self.create_table(name)
                csv_name = os.path.join(file_path, self.cm_truth + " " + name)
                Juice.save_matrix_table(self.matrix_tables[name], csv_name)
        return
    
    def _get_name_and_rank(self):
        Chai = cm.comp.pp.Parser()
        truth_other = Chai.main(os.path.join(self.input_path, self.cm_truth), 1)
        tp = {}
        fn = {}
        fp = {}
        tn = {}
        skipped_tax_id = []
        
        for sample_num in truth_other:
            for m in self.matrix_dict:
                for tax_id in self.matrix_dict[m]:
                    if tax_id in truth_other[sample_num]:
                        tp[tax_id] = {}
                        tp[tax_id]["rank"] = truth_other[sample_num][tax_id][0]
                        tp[tax_id]["name"] = truth_other[sample_num][tax_id][1]
                        fn[tax_id] = {}
                        fn[tax_id]["rank"] = truth_other[sample_num][tax_id][0]
                        fn[tax_id]["name"] = truth_other[sample_num][tax_id][1]
                        fp[tax_id] = {}
                        fp[tax_id]["rank"] = truth_other[sample_num][tax_id][0]
                        fp[tax_id]["name"] = truth_other[sample_num][tax_id][1]
                        tn[tax_id] = {}
                        tn[tax_id]["rank"] = truth_other[sample_num][tax_id][0]
                        tn[tax_id]["name"] = truth_other[sample_num][tax_id][1]
                        
                        self.true_taxid.append(tax_id)
                    else:
                        skipped_tax_id.append(tax_id)
        
        if len(skipped_tax_id) > 0:
            for name in self.matrix_dict:
                temp_other = Chai.main(os.path.join(self.input_path, name), 1)
                for sample_num in temp_other:
                    for tax_id in skipped_tax_id:
                        if tax_id in temp_other[sample_num]:
                            tp[tax_id] = {}
                            tp[tax_id]["rank"] = temp_other[sample_num][tax_id][0]
                            tp[tax_id]["name"] = temp_other[sample_num][tax_id][1]
                            fn[tax_id] = {}
                            fn[tax_id]["rank"] = temp_other[sample_num][tax_id][0]
                            fn[tax_id]["name"] = temp_other[sample_num][tax_id][1]
                            fp[tax_id] = {}
                            fp[tax_id]["rank"] = temp_other[sample_num][tax_id][0]
                            fp[tax_id]["name"] = temp_other[sample_num][tax_id][1]
                            tn[tax_id] = {}
                            tn[tax_id]["rank"] = temp_other[sample_num][tax_id][0]
                            tn[tax_id]["name"] = temp_other[sample_num][tax_id][1]
        return tp, fn, fp, tn
    
    def _get_true_positives(self, TP):
        for m in self.matrix_dict:
            for tax_id in self.matrix_dict[m]:
                if tax_id not in TP:
                    TP[tax_id] = {}
                TP[tax_id][m] = self.matrix_dict[m][tax_id][0]
        return TP
    def _get_false_negatives(self, FN):
        for m in self.matrix_dict:
            for tax_id in self.matrix_dict[m]:
                if tax_id not in FN:
                    FN[tax_id] = {}
                FN[tax_id][m] = self.matrix_dict[m][tax_id][1]
        return FN
    def _get_false_positives(self, FP):
        for m in self.matrix_dict:
            for tax_id in self.matrix_dict[m]:
                if tax_id not in FP:
                    FP[tax_id] = {}
                FP[tax_id][m] = self.matrix_dict[m][tax_id][2]
        return FP
    def _get_true_negatives(self, TN):
        for m in self.matrix_dict:
            for tax_id in self.matrix_dict[m]:
                if tax_id not in TN:
                    TN[tax_id] = {}
                TN[tax_id][m] = self.matrix_dict[m][tax_id][3]
        return TN
    
    def _organize_matrix(self):
        Juice = cm.Confusion(os.path.join(self.input_path, self.cm_truth), "")
        tp, fn, fp, tn = self._get_name_and_rank()
        
        names = self.get_matrix_names()
        TP = self._get_true_positives(tp)
        FN = self._get_false_negatives(fn)
        FP = self._get_false_positives(fp)
        TN = self._get_true_negatives(tn)
        
        all_tax_ids = set(TP.keys()) | set(FN.keys()) | set(FP.keys()) | set(TN.keys())
        matrix_sum = Juice.matrix_sum()
        
        for tax_id in all_tax_ids:
            for name in names:
                if (name not in TP[tax_id]) and (name not in FN[tax_id]) and (name not in FP[tax_id]) and (name not in TN[tax_id]):
                    TN[tax_id][name] = matrix_sum
                
                if (name not in TP[tax_id]):
                    TP[tax_id][name] = 0
                if name not in FN[tax_id]:
                    FN[tax_id][name] = 0
                if name not in FP[tax_id]:
                    FP[tax_id][name] = 0
        return TP, FN, FP, TN
    
    def organize_matrix(self):
        Calc = Precall()
        tp, fn, fp, tn = self._organize_matrix()
        precision, recall = Calc.calculate_precision_and_recall(self.matrix_dict)
        
        #### Creating DataFrames
        tp_df = pd.DataFrame.from_dict(tp, orient='index')
        fn_df = pd.DataFrame.from_dict(fn, orient="index")
        fp_df = pd.DataFrame.from_dict(fp, orient="index")
        tn_df = pd.DataFrame.from_dict(tn, orient="index")
        precision_df = pd.DataFrame.from_dict(precision, orient='index')
        recall_df = pd.DataFrame.from_dict(recall, orient="index")
        
        ### Creating 'Aggregate' column
        tp_df['Aggregate'] = tp_df.sum(axis=1)
        fn_df['Aggregate'] = fn_df.sum(axis=1)
        fp_df['Aggregate'] = fp_df.sum(axis=1)
        tn_df['Aggregate'] = tn_df.sum(axis=1)
        precision_df['Aggregate'] = tp_df['Aggregate'] / (tp_df['Aggregate'] + fp_df['Aggregate']) # TP / (TP+FP)
        recall_df['Aggregate'] = tp_df['Aggregate'] / (tp_df['Aggregate'] + fn_df['Aggregate']) # TP / (TP+FN)
        
        ### Sorting
        tp_df.sort_index(inplace=True)
        fn_df.sort_index(inplace=True)
        fp_df.sort_index(inplace=True)
        tn_df.sort_index(inplace=True)
        precision_df.sort_index(inplace=True)
        recall_df.sort_index(inplace=True)
        
        ### Changing np.nan to 'nan'
        precision_df = precision_df.fillna('nan')
        recall_df = recall_df.fillna('nan')
        return tp_df, fn_df, fp_df, tn_df, precision_df, recall_df
    
    def _write_col_title(self, path):
        workbook = load_workbook(path)
        for name in workbook.sheetnames:
            sheet = workbook[name]
            c = sheet["A1"]
            c.value = "Tax ID"
            c.font = Font(bold=True)
            c.alignment = Alignment(horizontal='center', vertical='center')
        workbook.save(path)
        print("\nSaved as \'{}\'".format(path))
        return
    def save_as_excel(self, file_path, file_name):
        tp, fn, fp, tn, precision, recall = self.organize_matrix()
        
        excel_name = os.path.join(file_path, file_name + ".xlsx")
        
        with pd.ExcelWriter(excel_name) as writer:
            tp.to_excel(writer, sheet_name="True Positives")
            fn.to_excel(writer, sheet_name="False Negatives")
            fp.to_excel(writer, sheet_name="False Positives")
            tn.to_excel(writer, sheet_name="True Negatives")
            precision.to_excel(writer, sheet_name="Precision")
            recall.to_excel(writer, sheet_name="Recall")
        
        self._write_col_title(excel_name)
        return
    
    
    def main(self, gnd_truth, excel_name="TaxaPerformanceMetrics_byTool", gen_dir="", file_path="", csv="no"):
        gen_paths = glob(os.path.join(gen_dir, "*.profile"))
        self.input_path = gen_dir
        self.output_path = file_path
        self.output_name = excel_name
        
        Juice = cm.Confusion(os.path.join(self.input_path, gnd_truth), "")
        self.set_truth(gnd_truth)
        
        for path in gen_paths:
            name = os.path.basename(path)
            if (name != gnd_truth) and (name not in self.matrix_dict):
                Juice.set_file_name(path)
                self.add_matrix(name, Juice.main("no"))
        
        if csv.lower() == "yes":
            self.save_matrices_as_csv(self.output_path)
        
        self.save_as_excel(self.output_path, excel_name)
        
        
        sheets = ["True Positives", "False Negatives", "False Positives", "True Negatives"]
        for sheet in sheets:
            ranks = self.read_excel(sheet, os.path.join(self.output_path, excel_name + ".xlsx"))
            ranks.append("")
            for rank in ranks:
                self.create_dendrogram(sheet, rank, os.path.join(self.output_path, excel_name + ".xlsx"))
        print("\nThe Dendrograms have been saved in {}.".format(self.output_path))
        
        return
    
    def read_excel(self,sheet, excel_path):
        ranks = []
        excel_df = pd.read_excel(excel_path, sheet_name=sheet)
        for rank in excel_df.loc[:, 'rank']:
            if rank not in ranks:
                ranks.append(rank)
        return ranks
    
    def create_dendrogram(self, metric, rank, excel_path):
        to_remove = ['Tax ID', 'rank', 'name', 'Aggregate']
        cols = []
        
        if rank == '':
            # make dendrogram over all ranks
            tmp_df = self.trace_back(metric)
        else:
            df = pd.read_excel(excel_path, sheet_name=metric)
            tmp_df = df[df['rank'] == rank]
            cols = [col for col in df.columns if col not in to_remove]
        
        
        tool_array = []
        names = []
        for item in cols:
            res = tmp_df[item]
            if np.sum(res) == 0:
                continue
            tool_array.append(res.tolist())
            names.append(item.split('.')[0])
        tool_array = np.array(tool_array)
        
        if len(tool_array) > 1:
            bray_curt = distance.pdist(np.array(tool_array), 'braycurtis')
            
            link = linkage(bray_curt, 'average')
            set_link_color_palette(['y', 'c', 'g', 'm', 'r'])
            
            plt.figure(figsize=[20.4, 10.4],dpi=480)
            title = metric + ": " + rank + "-Dendrogram"
            plt.suptitle(title, size=36, weight='semibold')
            den = dendrogram(link, orientation='right', labels=names)
            
            plt.xlim(-0.1, 1.1)
            fn = title.replace(": ", "-")
            filename = fn.replace(" ", "_") + '.png'
            plt.savefig(os.path.join(self.output_path, filename), dpi=480, facecolor='#B4FFDC', transparent=False, bbox_inches='tight')
            
            plt.close()
            print("\n{} has been saved.".format(filename))
        #plt.show()
        
        # add arg to create subplot grouped by metric or rank (subplot='none'; 'metric'; 'rank')
        return
    
    def trace_back(self, metric):
        Chai = cm.comp.pp.Parser()
        true_samples = Chai.main(os.path.join(self.input_path, self.cm_truth), 1)
        preds = self.get_matrix_names()
        true_data = {}
        data = {}
        
        # to get true data
        sample_sum = 0
        for sample in true_samples:
            sample_sum += 1
            true_data[sample] = set()
            for tax_id in true_samples[sample]:
                parent_ids = re.split('\|', true_samples[sample][tax_id][2])
                for parent in parent_ids:
                    true_data[sample].add(int(parent.strip()))
        
        # to get perdicted data
        for name in preds:
            data[name] = {}
            matrix = Chai.main(os.path.join(self.input_path, name), 1)
            for sample in matrix:
                data[name][sample] = set()
                for tax_id in matrix[sample]:
                    parent_ids = re.split('\|', matrix[sample][tax_id][2])
                    for parent in parent_ids:
                        data[name][sample].add(int(parent.strip()))
        
        Tea = cm.comp.Comparator()
        Juice = cm.Confusion('', '')
        new_matrix = {}
        for name in data:
            combined_taxid = Tea.combine_tax_ID(true_data, data[name])
            new_matrix[name] = Juice.confusion_matrix(true_data, data[name], combined_taxid)
        
        # make a data from for the correct metric
        df = pd.DataFrame()
        if metric == 'True Postives':
            TP = {}
            for name in new_matrix:
                for tax_id in new_matrix[name]:
                    if tax_id not in TP:
                        TP[tax_id] = {}
                    if name not in TP[tax_id]:
                        TP[tax_id][name] = new_matrix[name][tax_id][0]
            df = pd.DataFrame.from_dict(TP, orient='index')
        elif metric == 'False Negatives':
            FN = {}
            for name in new_matrix:
                for tax_id in new_matrix[name]:
                    if tax_id not in FN:
                        FN[tax_id] = {}
                    if name not in FN[tax_id]:
                        FN[tax_id][name] = new_matrix[name][tax_id][1]
            df = pd.DataFrame.from_dict(FN, orient='index')
        elif metric == 'False Postives':
            FP = {}
            for name in new_matrix:
                for tax_id in new_matrix[name]:
                    if tax_id not in FP:
                        FP[tax_id] = {}
                    if name not in FP[tax_id]:
                        FP[tax_id][name] = new_matrix[name][tax_id][2]
            df = pd.DataFrame.from_dict(FP, orient='index')
        elif metric == 'True Negatives':
            TN = {}
            for name in new_matrix:
                for tax_id in new_matrix[name]:
                    if tax_id not in TN:
                        TN[tax_id] = {}
                    if name not in TN[tax_id]:
                        TN[tax_id][name] = new_matrix[name][tax_id][3]
            df = pd.DataFrame.from_dict(TN, orient='index')
            
        
        return df.fillna(0)
    
    def get_top_taxid(self, x, metric='precall', difficulty='easy', truth="no"):
        excel_name =os.path.join(self.output_path, self.output_name) + '.xlsx'
        metric_df = pd.DataFrame()
        
        metric_df['Tax ID'] = pd.read_excel(excel_name, sheet_name='Precision')['Tax ID']
        if metric.lower() == 'precall':
            metric_df['Pre-Agg'] = pd.read_excel(excel_name, sheet_name='Precision')['Aggregate']
            metric_df['Re-Agg'] = pd.read_excel(excel_name, sheet_name='Recall')['Aggregate']
            metric_df['Average'] = (metric_df['Pre-Agg'] + metric_df['Re-Agg']) / 2
            base = 'Average'
        elif metric.lower() == 'tp':
            base = 'TP-Agg'
            metric_df[base] = pd.read_excel(excel_name, sheet_name='True Positives')['Aggregate']
        elif metric.lower() == 'fn':
            base = 'FN-Agg'
            metric_df[base] = pd.read_excel(excel_name, sheet_name='False Negatives')['Aggregate']
        elif metric.lower() == 'fp':
            base = 'FP-Agg'
            metric_df[base] = pd.read_excel(excel_name, sheet_name='False Positives')['Aggregate']
        elif metric.lower() == 'tn':
            base = 'TN-Agg'
            metric_df[base] = pd.read_excel(excel_name, sheet_name='True Negatives')['Aggregate']
        
        if truth.lower() == 'yes':
            Juice = cm.Confusion(self.cm_truth, '')
            Tea = cm.comp.Comparator()
            Chai = cm.comp.pp.Parser()
            untrue_taxids = Juice.dictionary_to_set(Tea.save_tax_ID(Chai.main(os.path.join(self.input_path, self.cm_truth)))) ^ set(metric_df['Tax ID'])
            untrue_indices = []
            for utt in untrue_taxids:
                untrue_indices.append(metric_df[metric_df['Tax ID']==utt].index.values[0])
            metric_df.drop(untrue_indices, inplace=True)
        
        if difficulty.lower() == 'easy':
            order = False
            nan_pos = 'last'
        elif difficulty.lower() == 'hard':
            order = True
            nan_pos = 'last'
        elif difficulty.lower() == 'nan':
            order = True
            nan_pos = 'first'
            
        return metric_df.sort_values(by=base, ascending=order, na_position=nan_pos).iloc[0:x, :]


#%% MAIN

if __name__ == "__main__":
    '''
    Choco = Misc()
    
    Choco.main(["truth", "pred", "pred2", "pred3"], "C:\\Users\\milkg\\Documents\\", "Derp_Test", "yes")
    '''
    
